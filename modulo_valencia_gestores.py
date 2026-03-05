from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook


def _normalizar(texto: str) -> str:
    return " ".join(str(texto).strip().split())


def generar_libros_gestores(
    ruta_excel_final: str,
    ruta_asignacion: str,
    carpeta_salida: str
) -> dict:

    resultado = {
        "ok": False,
        "errores": [],
        "archivos_generados": {}
    }

    try:

        # -------------------------------------------------
        # Validaciones básicas
        # -------------------------------------------------
        if not Path(ruta_excel_final).exists():
            resultado["errores"].append("No existe el Excel final validado.")
            return resultado

        if not Path(ruta_asignacion).exists():
            resultado["errores"].append("No existe el archivo gestor_zonas.xlsx.")
            return resultado

        Path(carpeta_salida).mkdir(parents=True, exist_ok=True)

        fecha_hoy = datetime.today().strftime("%Y-%m-%d")

        # -------------------------------------------------
        # Detectar hojas territoriales
        # -------------------------------------------------
        xls = pd.ExcelFile(ruta_excel_final)

        hojas_libro = xls.sheet_names
        zonas_libro_raw = [h for h in hojas_libro if h.startswith("ZREP_")]

        if not zonas_libro_raw:
            resultado["errores"].append(
                "No se han encontrado hojas territoriales (ZREP_*) en el libro final."
            )
            return resultado

        mapa_hojas = {_normalizar(z): z for z in zonas_libro_raw}
        zonas_libro_set = set(mapa_hojas.keys())

        # -------------------------------------------------
        # Leer asignación
        # -------------------------------------------------
        df_asignacion = pd.read_excel(ruta_asignacion)

        if not {"ZONA_REP", "GESTOR"}.issubset(df_asignacion.columns):
            resultado["errores"].append(
                "El archivo gestor_zonas.xlsx debe contener columnas ZONA_REP y GESTOR."
            )
            return resultado

        df_asignacion["ZONA_REP"] = df_asignacion["ZONA_REP"].apply(_normalizar)
        df_asignacion["GESTOR"] = df_asignacion["GESTOR"].astype(str).str.strip()

        duplicadas = df_asignacion["ZONA_REP"][df_asignacion["ZONA_REP"].duplicated()]
        if not duplicadas.empty:
            resultado["errores"].append(
                f"Zonas duplicadas en gestor_zonas.xlsx: {list(duplicadas)}"
            )
            return resultado

        zonas_asignadas = set(df_asignacion["ZONA_REP"])
        mapa_zona_gestor = dict(zip(df_asignacion["ZONA_REP"], df_asignacion["GESTOR"]))
        gestores_detectados = sorted(df_asignacion["GESTOR"].unique())

        # -------------------------------------------------
        # Validación crítica
        # -------------------------------------------------
        zonas_sin_gestor = zonas_libro_set - zonas_asignadas

        if zonas_sin_gestor:
            resultado["errores"].append(
                f"Zonas sin asignación en gestor_zonas.xlsx: {list(zonas_sin_gestor)}"
            )
            return resultado

        # -------------------------------------------------
        # Generación por gestor
        # -------------------------------------------------
        for gestor in gestores_detectados:

            zonas_gestor_norm = [
                z for z in zonas_libro_set if mapa_zona_gestor[z] == gestor
            ]

            if not zonas_gestor_norm:
                continue

            wb = Workbook()
            wb.remove(wb.active)

            dfs_todo = []

            for zona_norm in zonas_gestor_norm:

                zona_real = mapa_hojas[zona_norm]

                df_zona = pd.read_excel(ruta_excel_final, sheet_name=zona_real)

                ws = wb.create_sheet(title=zona_real)

                ws.append(list(df_zona.columns))

                for fila in df_zona.itertuples(index=False):
                    ws.append(list(fila))

                df_zona["ZONA"] = zona_real
                dfs_todo.append(df_zona)

            df_todo = pd.concat(dfs_todo, ignore_index=True)

            # -------------------------------------------------
            # Hoja TODO
            # -------------------------------------------------
            ws_todo = wb.create_sheet("TODO")

            ws_todo.append(list(df_todo.columns))

            for fila in df_todo.itertuples(index=False):
                ws_todo.append(list(fila))

            # -------------------------------------------------
            # RESUMEN_UNICO
            # -------------------------------------------------
            
            ws_resumen = wb.create_sheet("RESUMEN_UNICO")
            
            # mover hoja al inicio del libro
            wb._sheets.remove(ws_resumen)
            wb._sheets.insert(0, ws_resumen)
            
            ws_resumen["A1"] = "Total expediciones"
            ws_resumen["B1"] = "=COUNTA(TODO!A:A)-1"
            
            if "Kgs" in df_todo.columns:
            
                col_kgs = df_todo.columns.get_loc("Kgs") + 1
                col_letter = ws_todo.cell(row=1, column=col_kgs).column_letter
            
                ws_resumen["A2"] = "Total Kgs"
                ws_resumen["B2"] = f"=SUM(TODO!{col_letter}:{col_letter})"
            
            ws_resumen["A4"] = "Totales por zona"
            
            col_zona = df_todo.columns.get_loc("ZONA") + 1
            col_zona_letter = ws_todo.cell(row=1, column=col_zona).column_letter
            
            zonas_unicas = sorted(df_todo["ZONA"].unique())
            
            fila_inicio = 5
            
            for i, zona in enumerate(zonas_unicas):
            
                fila = fila_inicio + i
            
                ws_resumen[f"A{fila}"] = zona
                ws_resumen[f"B{fila}"] = (
                    f'=COUNTIF(TODO!{col_zona_letter}:{col_zona_letter},"{zona}")'
                )

            # -------------------------------------------------
            # Guardar archivo
            # -------------------------------------------------
            nombre_archivo = f"VALENCIA_{fecha_hoy}_{gestor}.xlsx"

            ruta_salida = Path(carpeta_salida) / nombre_archivo

            wb.save(ruta_salida)

            resultado["archivos_generados"][gestor] = str(ruta_salida)

        resultado["ok"] = True
        return resultado

    except Exception as e:
        resultado["errores"].append(str(e))
        return resultado
