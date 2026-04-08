from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import quote_sheetname


def encontrar_columna(ws, nombre):
    for row in range(1, min(10, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=row, column=col).value == nombre:
                return col
    return None


def generar_resumen_unico(ruta_excel: str, paradas_por_hoja: dict = None) -> None:

    wb = load_workbook(ruta_excel)

    if "RESUMEN_UNICO" in wb.sheetnames:
        del wb["RESUMEN_UNICO"]

    operativas = []
    if "ALMACEN" in wb.sheetnames:
        operativas.append("ALMACEN")
    if "HOSPITALES" in wb.sheetnames:
        operativas.append("HOSPITALES")
    if "FEDERACION" in wb.sheetnames:
        operativas.append("FEDERACION")
    zrep = sorted([s for s in wb.sheetnames if s.startswith("ZREP_")])
    operativas.extend(zrep)

    ws_res = wb.create_sheet("RESUMEN_UNICO", 0)
    ws_res.append(["Clave", "Expediciones", "Bultos", "Kilos", "Paradas"])
    for cell in ws_res[1]:
        cell.font = Font(bold=True)

    for fila_idx, hoja in enumerate(operativas, start=2):
        ws = wb[hoja]
        col_bultos = encontrar_columna(ws, "Bultos")
        col_kilos = encontrar_columna(ws, "Kgs") or encontrar_columna(ws, "Kilos")

        if col_bultos is None or col_kilos is None:
            continue

        letra_bultos = ws.cell(row=1, column=col_bultos).column_letter
        letra_kilos = ws.cell(row=1, column=col_kilos).column_letter
        paradas = paradas_por_hoja.get(hoja, "") if paradas_por_hoja else ""

        ws_res.append([
            hoja,
            f"=COUNTA('{hoja}'!A:A)-1",
            f"=SUM('{hoja}'!{letra_bultos}:{letra_bultos})",
            f"=SUM('{hoja}'!{letra_kilos}:{letra_kilos})",
            paradas
        ])

        # Hipervínculo en columna Clave → hoja correspondiente
        cell = ws_res.cell(row=fila_idx, column=1)
        cell.hyperlink = f"#{quote_sheetname(hoja)}!A1"
        cell.font = Font(color="0000FF", underline="single")

        # Hipervínculo de regreso en cada hoja → RESUMEN_UNICO en A1
        ws.insert_rows(1)
        cell_back = ws.cell(row=1, column=1)
        cell_back.value = "← RESUMEN"
        cell_back.hyperlink = f"#{quote_sheetname('RESUMEN_UNICO')}!A1"
        cell_back.font = Font(color="0000FF", underline="single", bold=True)

    ws_res.column_dimensions["A"].width = 30
    ws_res.column_dimensions["B"].width = 15
    ws_res.column_dimensions["C"].width = 15
    ws_res.column_dimensions["D"].width = 15
    ws_res.column_dimensions["E"].width = 12

    wb.save(ruta_excel)
