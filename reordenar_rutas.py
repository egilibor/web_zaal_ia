#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
from geocodificador import geocodificar
from openpyxl.styles import PatternFill
import pandas as pd
import re
import googlemaps
import requests
import datetime
import barcode
from barcode.writer import ImageWriter
from openpyxl.drawing.image import Image as XLImage
import io

#-----------------------------------------------------
# FUNCION PARA GENERAR CÓDIGO DE BARRAS
#-----------------------------------------------------

def generar_barcode_imagen(codigo: str) -> io.BytesIO:
    buffer = io.BytesIO()
    code128 = barcode.get("code128", str(codigo), writer=ImageWriter())
    code128.write(buffer, options={
        "module_height": 8,
        "module_width": 0.2,
        "font_size": 5,
        "text_distance": 2,
        "quiet_zone": 2,
    })
    buffer.seek(0)
    return buffer
    
# -------------------------------------------------
# ORÍGENES
# -------------------------------------------------

LAT_CASTELLON = 39.804106
LON_CASTELLON = -0.217351

LAT_VALENCIA = 39.44069
LON_VALENCIA = -0.42589


# -------------------------------------------------
# COLUMNAS OBLIGATORIAS
# -------------------------------------------------

COLUMNAS_OBLIGATORIAS = [
    "Exp",
    "Hospital",
    "Población",
    "Dirección",
    "Consignatario",
    "Cliente",
    "Kgs",
    "Bultos",
    "Z.Rep",
    "N. servicio",
]

# -------------------------------------------------
# NORMALIZAR TEXTO
# -------------------------------------------------

def normalizar_texto(txt):

    if pd.isna(txt):
        return ""

    txt = str(txt).strip().upper()

    txt = (
        txt.replace("Á", "A")
        .replace("É", "E")
        .replace("Í", "I")
        .replace("Ó", "O")
        .replace("Ú", "U")
        .replace("Ü", "U")
        .replace("Ñ", "N")
    )

    txt = " ".join(txt.split())

    return txt


# -------------------------------------------------
# DISTANCIA EUCLIDIANA
# -------------------------------------------------

def distancia(a, b):
    if a[0] is None or a[1] is None or b[0] is None or b[1] is None:
        return float('inf')
    return (a[0] - b[0]) ** 2 + (a[1] - b[1]) ** 2


# -------------------------------------------------
# PARADAS
# -------------------------------------------------

def extraer_calle_sin_numero(direccion: str) -> str:
    if pd.isna(direccion):
        return ""
    m = re.match(r"(.*?)[,\s]+\d+.*$", str(direccion).strip())
    if m:
        return m.group(1).strip()
    return str(direccion).strip()


def calcular_paradas_por_hoja(hojas_resultado: dict) -> dict:
    UMBRAL = 0.0009  # ~100 metros
    paradas = {}
    for nombre, df in hojas_resultado.items():
        if nombre in ("RESUMEN_UNICO", "METADATOS"):
            continue
        if "Latitud" not in df.columns or "Longitud" not in df.columns:
            continue

        coords_validas = []
        for _, row in df.iterrows():
            lat = row.get("Latitud")
            lon = row.get("Longitud")
            if lat is not None and lon is not None and pd.notna(lat) and pd.notna(lon):
                coords_validas.append((float(lat), float(lon)))

        paradas_unicas = []
        for coord in coords_validas:
            es_nueva = True
            for p in paradas_unicas:
                if abs(coord[0] - p[0]) <= UMBRAL and abs(coord[1] - p[1]) <= UMBRAL:
                    es_nueva = False
                    break
            if es_nueva:
                paradas_unicas.append(coord)

        paradas[nombre] = len(paradas_unicas)

    return paradas


# -------------------------------------------------
# 2-OPT (se mantiene como utilidad auxiliar)
# -------------------------------------------------

def mejorar_ruta_2opt(coords):

    mejor = coords.copy()
    mejora = True

    while mejora:

        mejora = False

        for i in range(1, len(mejor) - 2):
            for j in range(i + 1, len(mejor)):

                if j - i == 1:
                    continue

                a = mejor[i - 1]
                b = mejor[i]
                c = mejor[j - 1]
                d = mejor[j % len(mejor)]

                actual = distancia(a, b) + distancia(c, d)
                nuevo = distancia(a, c) + distancia(b, d)

                if nuevo < actual:
                    mejor[i:j] = reversed(mejor[i:j])
                    mejora = True

    return mejor


# -------------------------------------------------
# GOOGLE MAPS LINKS
# -------------------------------------------------

def generar_link_pueblos(df_ruta, lat_origen, lon_origen):

    puntos = [f"{lat_origen},{lon_origen}"]
    coords_vistas = set()

    for _, row in df_ruta.iterrows():
        lat = row.get("Latitud")
        lon = row.get("Longitud")
        if pd.notna(lat) and pd.notna(lon):
            clave = (round(float(lat), 5), round(float(lon), 5))
            if clave not in coords_vistas:
                coords_vistas.add(clave)
                puntos.append(f"{clave[0]},{clave[1]}")

    if len(puntos) < 2:
        return ""

    return "https://www.google.com/maps/dir/" + "/".join(puntos)


def generar_links_segmentos(df_ruta, lat_origen, lon_origen, tamanio=9):

    puntos = []
    coords_vistas = set()

    for _, row in df_ruta.iterrows():
        lat = row.get("Latitud")
        lon = row.get("Longitud")
        if pd.notna(lat) and pd.notna(lon):
            clave = (round(float(lat), 5), round(float(lon), 5))
            if clave not in coords_vistas:
                coords_vistas.add(clave)
                puntos.append(f"{clave[0]},{clave[1]}")

    if not puntos:
        return []

    origen = f"{lat_origen},{lon_origen}"
    links = []

    for i in range(0, len(puntos), tamanio):
        segmento = puntos[i:i + tamanio]
        if i == 0:
            tramo = [origen] + segmento
        else:
            tramo = [puntos[i - 1]] + segmento
        links.append("https://www.google.com/maps/dir/" + "/".join(tramo))

    return links


# -------------------------------------------------
# CARGAR COORDENADAS
# -------------------------------------------------

def cargar_coordenadas(ruta):

    df = pd.read_excel(ruta)
    df.columns = df.columns.str.strip().str.upper()

    if not {"PUEBLO", "LATITUD", "LONGITUD"}.issubset(df.columns):
        raise ValueError(
            f"Columnas detectadas: {list(df.columns)}. "
            "Se esperaban: PUEBLO, LATITUD, LONGITUD."
        )

    coords = {}

    for _, row in df.iterrows():
        pueblo = normalizar_texto(row["PUEBLO"])
        lat = row["LATITUD"]
        lon = row["LONGITUD"]
        if pd.notna(pueblo) and pd.notna(lat) and pd.notna(lon):
            coords[pueblo] = (float(lat), float(lon))

    return coords


# -------------------------------------------------
# BUSCAR COORDENADAS DE REFERENCIA
# -------------------------------------------------

def buscar_coords_referencia(pueblo_norm, coords):
    """
    Busca coordenadas de referencia para validar la geocodificación.
    Primero búsqueda exacta, luego parcial.
    """
    if pueblo_norm in coords:
        return coords[pueblo_norm]

    for key in coords:
        if pueblo_norm in key or key in pueblo_norm:
            return coords[key]

    return None


# -------------------------------------------------
# REFERENCIA DE COORDENADAS POR CP
# -------------------------------------------------

def cargar_referencia_cp(delegacion: str) -> dict:
    """
    Carga el archivo de referencia correspondiente a la delegación y devuelve
    un diccionario {cp_str: (lat, lon)} con las coordenadas de cada CP.
    Valencia  → valencia_municipios_coordenadas.xlsx
    Castellón → Libro_de_Servicio_Castellon_con_coordenadas.xlsx
    """
    raiz = Path(__file__).resolve().parent
    if delegacion == "valencia":
        ruta = raiz / "valencia_municipios_coordenadas.xlsx"
        df = pd.read_excel(ruta)
    else:
        ruta = raiz / "Libro_de_Servicio_Castellon_con_coordenadas.xlsx"
        df = pd.read_excel(ruta)

    df.columns = df.columns.str.strip().str.upper()

    cp_coords = {}
    for _, row in df.iterrows():
        cp = str(row.get("CODPOS", "")).strip().zfill(5)
        lat = row.get("LATITUD")
        lon = row.get("LONGITUD")
        if cp and pd.notna(lat) and pd.notna(lon):
            if cp not in cp_coords:
                cp_coords[cp] = (float(lat), float(lon))

    return cp_coords


# -------------------------------------------------
# ORDENACIÓN EN BLOQUES (API + fallback euclidiano)
# -------------------------------------------------

def ordenar_en_bloques(origen, waypoints, api_key, MAX_WAYPOINTS=25):
    """
    Ordena waypoints con la Routes API en bloques de MAX_WAYPOINTS.
    Si hay más de MAX_WAYPOINTS puntos, hace un primer paso euclidiano para
    obtener un orden inicial y luego refina cada bloque de 25 con la API,
    encadenando el origen con el último punto del bloque anterior.
    Usa ordenar_euclidiano como fallback si la API falla.
    Devuelve lista de índices relativos a los waypoints de entrada.
    """
    if not waypoints:
        return []
    if len(waypoints) == 1:
        return [0]

    if len(waypoints) <= MAX_WAYPOINTS:
        if api_key and len(waypoints) >= 2:
            try:
                return ordenar_segmento_api(origen, waypoints, api_key)
            except Exception:
                pass
        return ordenar_euclidiano(origen, waypoints)

    # Pre-ordenar con euclídeo para agrupar puntos cercanos
    orden_eucl = ordenar_euclidiano(origen, waypoints)
    waypoints_preord = [waypoints[i] for i in orden_eucl]

    resultado = []
    orig_actual = origen

    for j in range(0, len(waypoints_preord), MAX_WAYPOINTS):
        sub = waypoints_preord[j : j + MAX_WAYPOINTS]
        if api_key and len(sub) >= 2:
            try:
                ord_sub = ordenar_segmento_api(orig_actual, sub, api_key)
            except Exception:
                ord_sub = ordenar_euclidiano(orig_actual, sub)
        else:
            ord_sub = list(range(len(sub)))

        for o in ord_sub:
            resultado.append(orden_eucl[j + o])
        orig_actual = sub[ord_sub[-1]]

    return resultado


# -------------------------------------------------
# ORDENACIÓN CON ROUTES API
# -------------------------------------------------

def ordenar_segmento_api(origen, waypoints_coords, api_key):
    try:
        url = "https://routes.googleapis.com/directions/v2:computeRoutes"

        headers = {
            "Content-Type": "application/json",
            "X-Goog-Api-Key": api_key,
            "X-Goog-FieldMask": "routes.optimizedIntermediateWaypointIndex"
        }

        body = {
            "origin": {
                "location": {"latLng": {"latitude": origen[0], "longitude": origen[1]}}
            },
            "destination": {
                "location": {"latLng": {"latitude": origen[0], "longitude": origen[1]}}
            },
            "intermediates": [
                {
                    "location": {"latLng": {"latitude": lat, "longitude": lon}}
                }
                for lat, lon in waypoints_coords
            ],
            "travelMode": "DRIVE",
            "optimizeWaypointOrder": True
        }

        r = requests.post(url, json=body, headers=headers, timeout=10)
        data = r.json()

        if "routes" in data and data["routes"]:
            orden = data["routes"][0].get("optimizedIntermediateWaypointIndex", [])
            return orden
        else:
            print(f"DEBUG Routes API sin resultado: {data}")

    except Exception as e:
        print(f"DEBUG Error Routes API: {e}")
        raise

    return list(range(len(waypoints_coords)))


def ordenar_euclidiano(origen, waypoints_coords):
    """Nearest-neighbor con distancia euclidiana como fallback."""
    restantes = list(range(len(waypoints_coords)))
    orden = []
    lat_actual, lon_actual = origen

    while restantes:
        dists = [
            (waypoints_coords[i][0] - lat_actual) ** 2 +
            (waypoints_coords[i][1] - lon_actual) ** 2
            for i in restantes
        ]
        min_idx = restantes[dists.index(min(dists))]
        orden.append(min_idx)
        lat_actual, lon_actual = waypoints_coords[min_idx]
        restantes.remove(min_idx)

    return orden


# -------------------------------------------------
# ORDENACIÓN ZREP
# -------------------------------------------------

def ordenar_dataframe_zrep(df, coords, lat_origen, lon_origen, api_key="", delegacion="castellon", hora_salida=None):

    for col in COLUMNAS_OBLIGATORIAS:
        if col not in df.columns:
            raise ValueError(f"Falta columna obligatoria: {col}")

    df = df.copy()
    df["Latitud"] = None
    df["Longitud"] = None

    # -------------------------------------------------
    # GEOCODIFICACIÓN
    # -------------------------------------------------
    for idx, row in df.iterrows():
        pueblo_norm = normalizar_texto(row["Población"])
        cp = str(row.get('C.P.', '')).strip()
        lat, lon = (None, None)

        if api_key:
            dir_limpia = str(row['Dirección']).strip()
            pob_limpia = str(row['Población']).strip()
            if dir_limpia.upper() not in ("NAN", "NONE", "") and pob_limpia.upper() not in ("NAN", "NONE", ""):
                provincia = "VALENCIA" if delegacion == "valencia" else "CASTELLON"
                if cp.upper() not in ("NAN", "NONE", ""):
                    direccion_completa = f"{dir_limpia}, {cp} {pob_limpia}, {provincia}, ESPAÑA"
                else:
                    direccion_completa = f"{dir_limpia}, {pob_limpia}, {provincia}, ESPAÑA"
                lat, lon = geocodificar(direccion_completa, api_key)

                # Validar proximidad al municipio esperado
                if lat is not None and lon is not None:
                    coords_ref = buscar_coords_referencia(pueblo_norm, coords)
                    if coords_ref is not None:
                        lat_ref, lon_ref = coords_ref
                        distancia_ref = ((lat - lat_ref) ** 2 + (lon - lon_ref) ** 2) ** 0.5
                        if distancia_ref > 0.1:
                            lat, lon = None, None

        # Fallback: coordenadas del municipio
        if (lat is None or lon is None):
            coords_ref = buscar_coords_referencia(pueblo_norm, coords)
            if coords_ref is not None:
                lat, lon = coords_ref

        if pd.notna(lat) and pd.notna(lon):
            df.at[idx, "Latitud"] = lat
            df.at[idx, "Longitud"] = lon

    # -------------------------------------------------
    # SEPARAR FILAS CON Y SIN COORDENADAS
    # -------------------------------------------------
    filas_con_coord = [
        (idx, float(df.at[idx, "Latitud"]), float(df.at[idx, "Longitud"]))
        for idx in df.index
        if pd.notna(df.at[idx, "Latitud"]) and pd.notna(df.at[idx, "Longitud"])
    ]
    filas_sin_coord = [
        idx for idx in df.index
        if pd.isna(df.at[idx, "Latitud"]) or pd.isna(df.at[idx, "Longitud"])
    ]

    if not filas_con_coord:
        return df

    # -------------------------------------------------
    # AGRUPAR EN PARADAS ÚNICAS POR PROXIMIDAD
    # -------------------------------------------------
    UMBRAL = 0.0009  # ~100 metros
    paradas_unicas = []
    idx_por_parada = []

    for idx, lat, lon in filas_con_coord:
        asignada = False
        for i, p in enumerate(paradas_unicas):
            if abs(lat - p[0]) <= UMBRAL and abs(lon - p[1]) <= UMBRAL:
                idx_por_parada[i].append(idx)
                asignada = True
                break
        if not asignada:
            paradas_unicas.append((lat, lon))
            idx_por_parada.append([idx])

    # -------------------------------------------------
    # AGRUPAR PARADAS POR C.P.
    # -------------------------------------------------
    # Asociar cada parada única con su C.P.
    cp_por_parada = []
    for i in range(len(paradas_unicas)):
        indices = idx_por_parada[i]
        cps = [
            str(df.at[idx, 'C.P.']).strip().zfill(5)
            for idx in indices
            if 'C.P.' in df.columns and pd.notna(df.at[idx, 'C.P.'])
        ]
        cp_por_parada.append(cps[0] if cps else '')

    grupos_cp = {}
    for i, cp in enumerate(cp_por_parada):
        grupos_cp.setdefault(cp, []).append(i)

    # -------------------------------------------------
    # PRIMERA PASADA — orden entre zonas (CPs)
    # -------------------------------------------------
    # Obtener coordenadas de referencia para cada CP desde el archivo de referencia
    try:
        cp_coords_ref = cargar_referencia_cp(delegacion)
    except Exception as e:
        print(f"DEBUG No se pudo cargar referencia CP: {e}")
        cp_coords_ref = {}

    lista_cps = list(grupos_cp.keys())
    centroides = []
    for cp in lista_cps:
        if cp in cp_coords_ref:
            centroides.append(cp_coords_ref[cp])
        else:
            # Fallback: centroide calculado de las paradas geocodificadas
            lats = [paradas_unicas[i][0] for i in grupos_cp[cp]]
            lons = [paradas_unicas[i][1] for i in grupos_cp[cp]]
            centroides.append((sum(lats) / len(lats), sum(lons) / len(lons)))

    # Ordenar CPs con la API (o euclídeo si no hay api_key)
    orden_cps_idx = ordenar_en_bloques((lat_origen, lon_origen), centroides, api_key)
    cps_ordenados = [lista_cps[i] for i in orden_cps_idx]

    # -------------------------------------------------
    # SEGUNDA PASADA — orden dentro de cada zona (CP)
    # -------------------------------------------------
    orden_paradas = []
    origen_actual = (lat_origen, lon_origen)

    for cp in cps_ordenados:
        indices_paradas_cp = grupos_cp[cp]
        coords_cp = [paradas_unicas[i] for i in indices_paradas_cp]

        orden_seg = ordenar_en_bloques(origen_actual, coords_cp, api_key)

        for o in orden_seg:
            orden_paradas.append(indices_paradas_cp[o])

        if orden_seg:
            origen_actual = coords_cp[orden_seg[-1]]

    # -------------------------------------------------
    # RECONSTRUIR ORDEN DE FILAS
    # -------------------------------------------------
    indices_ordenados = []
    for pos in orden_paradas:
        indices_ordenados.extend(idx_por_parada[pos])
    indices_ordenados.extend(filas_sin_coord)

    df_ordenado = df.loc[indices_ordenados].copy()

    # Agrupar por calle dentro de cada población sin alterar orden de poblaciones
    df_ordenado["Calle_sin_num"] = df_ordenado["Dirección"].apply(extraer_calle_sin_numero)
    df_ordenado["Clave_parada"] = (
        df_ordenado["Población"].astype(str).str.strip().str.upper()
        + "|"
        + df_ordenado["Calle_sin_num"].str.upper()
    )

    poblaciones_orden = {p: i for i, p in enumerate(df_ordenado["Población"].unique())}
    df_ordenado["_ord_pob"] = df_ordenado["Población"].map(poblaciones_orden)

    claves_orden = {c: i for i, c in enumerate(df_ordenado["Clave_parada"].unique())}
    df_ordenado["_ord_clave"] = df_ordenado["Clave_parada"].map(claves_orden)

    df_ordenado = df_ordenado.sort_values(["_ord_pob", "_ord_clave"])
    df_ordenado = df_ordenado.drop(columns=["Calle_sin_num", "Clave_parada", "_ord_pob", "_ord_clave"])

    return df_ordenado


# -------------------------------------------------
# FUNCIÓN PRINCIPAL
# -------------------------------------------------

def reordenar_excel(
    input_path: Path,
    output_path: Path,
    ruta_coordenadas: Path,
    lat_origen: float,
    lon_origen: float,
    api_key: str = "",
    delegacion: str = "castellon",
    hora_salida=None,
):

    hojas = pd.read_excel(input_path, sheet_name=None)
    coords = cargar_coordenadas(ruta_coordenadas)
    hojas_resultado = {}

    for nombre, df in hojas.items():

        if nombre.startswith("ZREP_") or nombre in ("HOSPITALES", "FEDERACION"):
            df_ordenado = ordenar_dataframe_zrep(
                df,
                coords,
                lat_origen,
                lon_origen,
                api_key=api_key,
                delegacion=delegacion,
                hora_salida=hora_salida,
            )

            link = generar_link_pueblos(df_ordenado, lat_origen, lon_origen)

            df_ordenado["NAVEGACIÓN"] = ""



            # Asignar número de parada por proximidad
            UMBRAL = 0.0009
            numeros_parada = []
            paradas_unicas = []

            for _, row in df_ordenado.iterrows():
                lat = row.get("Latitud")
                lon = row.get("Longitud")
                if lat is not None and lon is not None and pd.notna(lat) and pd.notna(lon):
                    asignada = False
                    for num, p in enumerate(paradas_unicas, 1):
                        if abs(float(lat) - p[0]) <= UMBRAL and abs(float(lon) - p[1]) <= UMBRAL:
                            numeros_parada.append(num)
                            asignada = True
                            break
                    if not asignada:
                        paradas_unicas.append((float(lat), float(lon)))
                        numeros_parada.append(len(paradas_unicas))
                else:
                    numeros_parada.append("")

            df_ordenado = df_ordenado.rename(columns={"Hospital": "Parada"})
            df_ordenado["Parada"] = numeros_parada
            hojas_resultado[nombre] = df_ordenado

        else:
            hojas_resultado[nombre] = df

    paradas_por_hoja = calcular_paradas_por_hoja(hojas_resultado)

    if "RESUMEN_UNICO" in hojas_resultado:
        df_res = hojas_resultado["RESUMEN_UNICO"].copy()
        if "Paradas" not in df_res.columns:
            df_res["Paradas"] = ""
        df_res["Paradas"] = df_res["Clave"].map(paradas_por_hoja).fillna(df_res["Paradas"])
        hojas_resultado["RESUMEN_UNICO"] = df_res

    # Construir enlaces de navegación por hoja
    hojas_navegacion = {}
    for nombre, df in hojas_resultado.items():
        if not nombre.startswith("ZREP_") and nombre not in ("HOSPITALES", "FEDERACION"):
            continue
        if "Latitud" not in df.columns or "Longitud" not in df.columns:
            continue
        link_completo = generar_link_pueblos(df, lat_origen, lon_origen)
        segmentos = generar_links_segmentos(df, lat_origen, lon_origen)
        hojas_navegacion[nombre] = {
            "link_completo": link_completo,
            "segmentos": segmentos
        }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for nombre, df in hojas_resultado.items():
            df.to_excel(writer, sheet_name=nombre, index=False)

    # Añadir filas de navegación al principio de cada hoja
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import quote_sheetname, get_column_letter

    wb = load_workbook(output_path)

    for nombre, datos in hojas_navegacion.items():
        if nombre not in wb.sheetnames:
            continue
        ws = wb[nombre]
        ws.insert_rows(1, amount=len(datos["segmentos"]) + 1)
        ws.cell(row=1, column=1).value = "RUTA COMPLETA"
        ws.cell(row=1, column=1).hyperlink = f"#{quote_sheetname('RESUMEN_UNICO')}!A1"
        ws.cell(row=1, column=2).value = datos["link_completo"]
        ws.cell(row=1, column=2).font = Font(color="0000FF", underline="single")
        for i, link in enumerate(datos["segmentos"]):
            ws.cell(row=2 + i, column=1).value = f"SEGMENTO {i + 1}"
            ws.cell(row=2 + i, column=2).value = link
            ws.cell(row=2 + i, column=2).font = Font(color="0000FF", underline="single")

        # Botón de regreso en la última fila de navegación (sin hipervínculo; el enlace ya está en A1)
        ultima_fila_nav = 1 + len(datos["segmentos"])
        cell_back = ws.cell(row=ultima_fila_nav, column=3)
        # cell_back.value = "← RESUMEN"
        cell_back.font = Font(color="FFFFFF", bold=True)
        #cell_back.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")

        # Eliminar hipervínculos en P1 y C2
        ws["P1"].hyperlink = None
        ws["C2"].hyperlink = None

        # Eliminar contenido de N4 (ruta Google Maps)
        ws.cell(row=len(datos["segmentos"]) + 3, column=15).value = None

    azul_claro = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")

    for nombre in hojas_navegacion.keys():
        if nombre not in wb.sheetnames:
            continue
        ws = wb[nombre]
        n_nav = len(hojas_navegacion[nombre]["segmentos"]) + 1

        for row in ws.iter_rows(min_row=n_nav + 2):
            valor_parada = row[1].value
            try:
                if int(valor_parada) % 2 != 0:
                    for cell in row:
                        cell.fill = azul_claro
            except (TypeError, ValueError):
                pass

    # ── Códigos de barras ──
    for nombre in hojas_navegacion.keys():
        if nombre not in wb.sheetnames:
            continue
        ws = wb[nombre]
        n_nav = len(hojas_navegacion[nombre]["segmentos"]) + 1

        # Insertar columna Barcode en columna C (posición 3), justo después de Parada (B)
        ws.insert_cols(3)
        col_barcode = 3
        ws.cell(row=n_nav + 1, column=col_barcode).value = "Barcode"
        ws.column_dimensions["C"].width = 16

        # Buscar columna "Exp" después del insert (los índices >= 3 se han desplazado)
        col_exp = None
        for cell in ws[n_nav + 1]:
            if cell.value == "Exp":
                col_exp = cell.column
                break

        if col_exp is None:
            continue

        for row_idx in range(n_nav + 2, ws.max_row + 1):
            exp_val = ws.cell(row=row_idx, column=col_exp).value
            if not exp_val:
                continue
            try:
                img_buffer = generar_barcode_imagen(str(exp_val))
                img = XLImage(img_buffer)
                img.width = 120
                img.height = 35
                celda = ws.cell(row=row_idx, column=col_barcode)
                ws.row_dimensions[row_idx].height = 28
                ws.add_image(img, celda.coordinate)
            except Exception:
                pass
                
    # ── Anchos de columna ──
    ANCHOS_COLUMNA = {"Exp": 15, "Población": 20, "Dirección": 40, "Consignatario": 35}
    for nombre in hojas_navegacion.keys():
        if nombre not in wb.sheetnames:
            continue
        ws = wb[nombre]
        n_nav = len(hojas_navegacion[nombre]["segmentos"]) + 1
        for cell in ws[n_nav + 1]:
            if cell.value in ANCHOS_COLUMNA:
                ws.column_dimensions[get_column_letter(cell.column)].width = ANCHOS_COLUMNA[cell.value]

    wb.save(output_path)
    return paradas_por_hoja
