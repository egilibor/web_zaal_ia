#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
ATRASOS · PENDIENTES (determinista)

- Entrada: CSV (pendientes no entregados)
- Corte: cierre de ayer (23:59:59) según fecha del sistema
- Salida: Excel "atrasos_YYYY-MM-DD.xlsx" en la misma carpeta del CSV
- Columnas (exactas):
  Exp, F.Llegada, Z.Rep, Consignatario, Población, Dir.Entrega, C.P., Días de atraso, Tramo
- Tramo (exacto): 24 / 48h / + de 48
- Días de atraso: entero (redondeo)
"""

from __future__ import annotations

import csv
import os
from datetime import datetime, date, time, timedelta

import numpy as np
import pandas as pd
from pandas.errors import ParserError
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment


REQUIRED_COLS = [
    "Exp",
    "F.Llegada",
    "Z.Rep",
    "Consignatario",
    "Población",
    "Dir. entrega",
    "C.P.",
]


def read_csv_robusto(path: str) -> pd.DataFrame:
    """
    Lectura robusta y determinista:
    1) Intenta autodetectar separador (engine=python) con UTF-8-SIG.
    2) Si hay ParserError por comillas rotas, fuerza sep=';' y desactiva quoting.
    3) Repite lo anterior con latin-1 si falla por encoding.
    """
    # Intento 1: autodetección
    try:
        return pd.read_csv(path, dtype=str, sep=None, engine="python", encoding="utf-8-sig")
    except UnicodeDecodeError:
        pass
    except ParserError:
        pass

    # Intento 2: fuerza ';' y desactiva quoting (típico en exports con comillas mal cerradas)
    try:
        return pd.read_csv(
            path,
            dtype=str,
            sep=";",
            engine="python",
            encoding="utf-8-sig",
            quoting=csv.QUOTE_NONE,
            escapechar="\\",
        )
    except UnicodeDecodeError:
        pass
    except ParserError:
        pass

    # Intento 3: latin-1 (último recurso)
    try:
        return pd.read_csv(path, dtype=str, sep=None, engine="python", encoding="latin-1")
    except ParserError:
        return pd.read_csv(
            path,
            dtype=str,
            sep=";",
            engine="python",
            encoding="latin-1",
            quoting=csv.QUOTE_NONE,
            escapechar="\\",
        )


def compute_cutoff_end_of_yesterday() -> datetime:
    today = date.today()
    yesterday = today - timedelta(days=1)
    return datetime.combine(yesterday, time(23, 59, 59))


def tramo_from_hours(h: float) -> str:
    # Etiquetas EXACTAS pedidas
    if pd.isna(h):
        return ""
    if h <= 24:
        return "24"
    if h <= 48:
        return "48h"
    return "+ de 48"


def build_excel(out_path: str, cutoff: datetime, out_df: pd.DataFrame) -> None:
    wb = Workbook()

    # PARAM
    ws_param = wb.active
    ws_param.title = "PARAM"
    ws_param["A1"] = "Corte (cierre de ayer)"
    ws_param["B1"] = cutoff
    ws_param["A1"].font = Font(bold=True)
    ws_param["B1"].number_format = "dd/mm/yyyy hh:mm:ss"
    ws_param.column_dimensions["A"].width = 28
    ws_param.column_dimensions["B"].width = 22

    # PENDIENTES
    ws = wb.create_sheet("PENDIENTES")
    headers = list(out_df.columns)

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(headers)
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.row_dimensions[1].height = 28

    # Volcado (con conversión NA->None para no romper openpyxl)
    for _, r in out_df.iterrows():
        row = []
        for h in headers:
            v = r[h]
            if pd.isna(v):
                row.append(None)
            elif h == "F.Llegada":
                row.append(v.to_pydatetime() if hasattr(v, "to_pydatetime") else v)
            else:
                row.append(v)
        ws.append(row)

    # Formatos y anchos
    widths = {
        "Exp": 12,
        "F.Llegada": 18,
        "Z.Rep": 14,
        "Consignatario": 34,
        "Población": 18,
        "Dir.Entrega": 42,
        "C.P.": 8,
        "Días de atraso": 14,
        "Tramo": 10,
    }

    for j, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(h, 14)
        if h == "F.Llegada":
            for i in range(2, ws.max_row + 1):
                ws.cell(i, j).number_format = "dd/mm/yyyy hh:mm"
        if h == "Días de atraso":
            for i in range(2, ws.max_row + 1):
                ws.cell(i, j).number_format = "0"

    wb.save(out_path)


def main() -> int:
    in_path = input("Nombre o ruta del CSV de entrada (pendientes): ").strip().strip('"').strip("'")
    if not in_path:
        print("ERROR: no se indicó archivo.")
        return 1

    if not os.path.isfile(in_path):
        print(f"ERROR: no existe el archivo: {in_path}")
        return 1

    df = read_csv_robusto(in_path)

    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        print("ERROR: faltan columnas obligatorias en el CSV:")
        for c in missing:
            print(f" - {c}")
        print("\nColumnas encontradas:", ", ".join(df.columns))
        return 1

    cutoff = compute_cutoff_end_of_yesterday()

    # Parseo fecha llegada (dayfirst=True)
    df["F.Llegada_dt"] = pd.to_datetime(df["F.Llegada"], dayfirst=True, errors="coerce")

    # Horas / días atraso
    hours = (cutoff - df["F.Llegada_dt"]) / pd.Timedelta(hours=1)
    days = (cutoff - df["F.Llegada_dt"]) / pd.Timedelta(days=1)

    # Días de atraso redondeados a ENTERO (manteniendo NA si falta fecha)
    df["Días de atraso"] = days.round(0).astype("Int64")

    # Tramo
    df["Tramo"] = hours.apply(tramo_from_hours)

    # Salida con campos EXACTOS y orden por mayor atraso
    out_df = pd.DataFrame({
        "Exp": df["Exp"],
        "F.Llegada": df["F.Llegada_dt"],
        "Z.Rep": df["Z.Rep"],
        "Consignatario": df["Consignatario"],
        "Población": df["Población"],
        "Dir.Entrega": df["Dir. entrega"],
        "C.P.": df["C.P."],
        "Días de atraso": df["Días de atraso"],
        "Tramo": df["Tramo"],
    }).sort_values(by="Días de atraso", ascending=False, na_position="last")

    out_name = f"atrasos_{date.today().isoformat()}.xlsx"
    out_path = os.path.join(os.path.dirname(os.path.abspath(in_path)), out_name)

    build_excel(out_path, cutoff, out_df)

    print(f"OK: generado {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
