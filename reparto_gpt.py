#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import os
import datetime as _dt
import re
from pathlib import Path
from typing import List

import pandas as pd
import difflib
import json
from geocodificador import geocodificar
from reordenar_rutas import cargar_coordenadas, buscar_coords_referencia, normalizar_texto
# -------------------------
# CALLEJERO CASTELLÓN
# -------------------------

CALLES_CASTELLON = []

try:
    calles_path = Path(__file__).parent / "calles_castellon.csv"
    if calles_path.exists():
        df_calles = pd.read_csv(calles_path)
        if "nombre" in df_calles.columns:
            CALLES_CASTELLON = (
                df_calles["nombre"]
                .astype(str)
                .str.upper()
                .unique()
                .tolist()
            )
except Exception:
    CALLES_CASTELLON = []


def corregir_calle_castellon(poblacion: str, direccion: str) -> str:

    if not direccion:
        return direccion

    if "CASTELL" not in norm(poblacion):
        return direccion

    if not CALLES_CASTELLON:
        return direccion

    # separar calle y número
    m = re.match(r"(.*?)[,\s]+(\d+.*)", direccion)
    if m:
        calle = m.group(1)
        numero = m.group(2)
    else:
        calle = direccion
        numero = ""

    calle_norm = norm(calle)

    # crear mapa normalizado → original
    mapa = {norm(c): c for c in CALLES_CASTELLON}

    match = difflib.get_close_matches(
        calle_norm,
        list(mapa.keys()),
        n=1,
        cutoff=0.85
    )

    if match:
        calle = mapa[match[0]]

    if numero:
        return f"{calle} {numero}"

    return calle

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# -------------------------
# UTILIDADES
# -------------------------

def clean_text(x) -> str:
    if pd.isna(x):
        return ""
    return re.sub(r"\s+", " ", str(x).strip())


_ILLEGAL_CHARS_RE = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]"
)

def sanitize_cell(x):
    """Elimina caracteres de control ilegales para openpyxl antes de escribir en Excel."""
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    if isinstance(x, str):
        return _ILLEGAL_CHARS_RE.sub("", x)
    return x


def parse_kg(x) -> float:
    if pd.isna(x):
        return 0.0
    s = str(x).replace(",", ".")
    s = re.sub(r"[^0-9\.\-]", "", s)
    try:
        return float(s)
    except:
        return 0.0


def parse_int(x) -> int:
    if pd.isna(x):
        return 0
    s = re.sub(r"[^0-9\-]", "", str(x))
    try:
        return int(s)
    except:
        return 0


def norm(s: str) -> str:
    s = clean_text(s).upper()
    trans = str.maketrans({
        "Á":"A","É":"E","Í":"I","Ó":"O","Ú":"U","Ü":"U","Ñ":"N","Ç":"C",
    })
    s = s.translate(trans)
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def extraer_calle_sin_numero(direccion: str) -> str:
    """Elimina el número final de una dirección para obtener solo la calle."""
    m = re.match(r"(.*?)[,\s]+\d+.*$", direccion.strip())
    if m:
        return m.group(1).strip()
    return direccion.strip()
    
def sheet_to_df(wb, name: str) -> pd.DataFrame:
    if name not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[name]
    data = list(ws.values)
    if not data:
        return pd.DataFrame()
    headers = list(data[0])
    rows = data[1:]
    return pd.DataFrame(rows, columns=headers)


def prepare_rules(df_rules: pd.DataFrame, pob_col: str, pat_col: str) -> pd.DataFrame:
    if df_rules.empty:
        return df_rules
    d = df_rules.copy()
    d["Pob_norm"] = d[pob_col].astype(str).apply(norm)
    d["Pat_norm"] = d[pat_col].astype(str).apply(norm)
    d = d[d["Pat_norm"].ne("")]
    return d


def match_rules(pob_norm: str, dir_norm: str, rules_df: pd.DataFrame, tag_field: str | None = None):
    for _, r in rules_df.iterrows():
        if r["Pob_norm"] and r["Pob_norm"] not in pob_norm:
            continue
        if r["Pat_norm"] and r["Pat_norm"] in dir_norm:
            if tag_field:
                return True, str(r.get(tag_field, "") or "")
            return True, ""
    return False, ""


def style_sheet(ws):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border

    ws.freeze_panes = "A2"


# -------------------------
# CORE
# -------------------------

def run(csv_path: Path, reglas_path: Path, out_path: Path, origen: str, delegacion: str,
        api_key: str = "", ruta_coordenadas: Path | None = None) -> None:

    try:
        df = pd.read_csv(
            csv_path,
            sep=";",
            encoding="utf-8-sig",
            dtype=str,
            engine="python",
            on_bad_lines="skip"
        )
    except UnicodeDecodeError:
        df = pd.read_csv(
            csv_path,
            sep=";",
            encoding="latin-1",
            dtype=str,
            engine="python",
            on_bad_lines="skip"
        )

    COL_MAP = {
        "N.Exp":    "Exp",
        "Cod.Pos":  "C.P.",
        "Domicilio": "Dir. entrega",
    }
    df.rename(columns={k: v for k, v in COL_MAP.items() if k in df.columns}, inplace=True)

    df["Exp"] = df["Exp"].astype(str).str.strip()

    if "Kgs" not in df.columns and "K.Doc" in df.columns:
        df["Kgs"] = df["K.Doc"]
    if "Btos." not in df.columns and "K.Doc" in df.columns:
        df["Btos."] = df["K.Doc"]

    for col in ["N. servicio", "Cliente"]:
        if col not in df.columns:
            df[col] = ""

    df["Kgs"] = df["Kgs"].apply(parse_kg)
    if "B.Doc" in df.columns:
        df["Bultos"] = df["B.Doc"].apply(parse_int)
    else:
        df["Bultos"] = df["Btos."].apply(parse_int)
    df["Población"] = df["Población"].fillna("")
    df["Dirección"] = df["Dir. entrega"].fillna("")
    df["Z.Rep"] = df["Z.Rep"].fillna("")
    df["Cliente"] = df.get("Cliente", "")

    # --- LIMPIEZA DIRECCIONES ---
    df["Dirección"] = df["Dirección"].apply(clean_text).str.upper()
    
    df["Dirección"] = df.apply(
        lambda r: corregir_calle_castellon(r["Población"], r["Dirección"]),
        axis=1
    )

    # -------------------------
    # GEOCODIFICACIÓN (Fase 1)
    # -------------------------
    df["Latitud"] = None
    df["Longitud"] = None

    coords_municipios = {}
    if ruta_coordenadas is not None:
        try:
            coords_municipios = cargar_coordenadas(ruta_coordenadas)
        except Exception as e:
            print(f"Aviso: no se pudo cargar coordenadas de municipios: {e}")

    if api_key or coords_municipios:
        provincia = "VALENCIA" if delegacion == "valencia" else "CASTELLON"
        for idx, row in df.iterrows():
            dir_limpia = str(row["Dirección"]).strip()
            pob_limpia = str(row["Población"]).strip()
            cp = str(row.get("C.P.", "") or "").strip()
            lat, lon = None, None

            if api_key and dir_limpia.upper() not in ("NAN", "NONE", "") and pob_limpia.upper() not in ("NAN", "NONE", ""):
                if cp.upper() not in ("NAN", "NONE", ""):
                    direccion_completa = f"{dir_limpia}, {cp} {pob_limpia}, {provincia}, ESPAÑA"
                else:
                    direccion_completa = f"{dir_limpia}, {pob_limpia}, {provincia}, ESPAÑA"
                lat, lon = geocodificar(direccion_completa, api_key)

                if lat is not None and lon is not None and coords_municipios:
                    pueblo_norm = normalizar_texto(pob_limpia)
                    coords_ref = buscar_coords_referencia(pueblo_norm, coords_municipios)
                    if coords_ref is not None:
                        lat_ref, lon_ref = coords_ref
                        if ((lat - lat_ref) ** 2 + (lon - lon_ref) ** 2) ** 0.5 > 0.1:
                            lat, lon = None, None

            # Fallback a coordenadas del municipio
            if (lat is None or lon is None) and coords_municipios:
                pueblo_norm = normalizar_texto(pob_limpia)
                coords_ref = buscar_coords_referencia(pueblo_norm, coords_municipios)
                if coords_ref is not None:
                    lat, lon = coords_ref

            if lat is not None and lon is not None:
                df.at[idx, "Latitud"] = lat
                df.at[idx, "Longitud"] = lon

    # -------------------------
    # APLICAR REGLAS
    # -------------------------

    df["Pob_norm"] = df["Población"].apply(norm)
    df["Dir_norm"] = df["Dirección"].apply(norm)

    wb_rules = load_workbook(reglas_path, data_only=True)
    rules_h = sheet_to_df(wb_rules, "REGLAS_HOSPITALES")
    rules_f = sheet_to_df(wb_rules, "REGLAS_FEDERACION")

    rules_h_prep = prepare_rules(rules_h, "Población", "Patrón_dirección")
    rules_f_prep = prepare_rules(rules_f, "Población", "Patrón_dirección")

    df["is_hospital"] = False
    df["Hospital"] = ""

    for i, r in df.iterrows():
        ok, tag = match_rules(r["Pob_norm"], r["Dir_norm"], rules_h_prep, "Hospital_final")
        if ok:
            df.at[i, "is_hospital"] = True
            df.at[i, "Hospital"] = tag

    df["is_fed"] = False
    for i, r in df.iterrows():
        ok, _ = match_rules(r["Pob_norm"], r["Dir_norm"], rules_f_prep)
        if ok:
            df.at[i, "is_fed"] = True

    df["is_any_special"] = df["is_hospital"] | df["is_fed"]
    df["Calle_sin_num"] = df["Dirección"].apply(extraer_calle_sin_numero)
    df["Clave_parada"] = df["Población"].str.strip().str.upper() + "|" + df["Calle_sin_num"].str.upper()
    hosp = df[df["is_hospital"]].copy()
    fed = df[df["is_fed"]].copy()
    resto = df[~df["is_any_special"]].copy()
    resto["Z.Rep"] = (
        resto["Z.Rep"]
        .astype(str)
        .str.strip()
        .replace(".", "")
    )
    
    # -------------------------
    # EXCEL
    # -------------------------

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    COLUMNAS_BASE = [
        "Exp", "Ref.", "Hospital", "Población", "Dirección",
        "Consignatario", "Cliente", "Kgs",
        "Bultos", "Z.Rep", "N. servicio", "C.P.",
        "Latitud", "Longitud",
    ]

    COLUMNAS_EXTRA = ["Remitente", "Tel.Contacto", "ObsClt", "F.Max.Ent", "Compromiso", "Prio.", "B.Doc", "F.Teo.Entr.", "Obs.", "AmpFtiI"]

    # METADATOS
    meta = pd.DataFrame({
        "Clave": ["Delegación", "Origen de datos", "CSV", "Reglas", "Generado"],
        "Valor": [
            delegacion,
            origen,
            str(csv_path),
            str(reglas_path),
            _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ],
    })

    ws_meta = wb_out.create_sheet("METADATOS")
    for row in dataframe_to_rows(meta, index=False, header=True):
        ws_meta.append([sanitize_cell(v) for v in row])
    style_sheet(ws_meta)

    # HOSPITALES
    cols_out = [c for c in COLUMNAS_BASE + COLUMNAS_EXTRA if c in df.columns]
    ws_h = wb_out.create_sheet("HOSPITALES")
    for row in dataframe_to_rows(hosp[cols_out], index=False, header=True):
        ws_h.append([sanitize_cell(v) for v in row])
    style_sheet(ws_h)

    # FEDERACION
    ws_f = wb_out.create_sheet("FEDERACION")
    for row in dataframe_to_rows(fed[cols_out], index=False, header=True):
        ws_f.append([sanitize_cell(v) for v in row])
    style_sheet(ws_f)

    # ZREP
    existing = set(wb_out.sheetnames)

    for z, sub in resto.groupby("Z.Rep"):

        z = str(z).strip()
        if z == ".":
            z = ""

        nombre = f"ZREP_{z}"
        nombre = re.sub(r"[\\/*?:\[\]]", "_", nombre)[:31]

        base = nombre
        i = 1
        while nombre in existing:
            sufijo = f"_{i}"
            nombre = (base[:31 - len(sufijo)] + sufijo)
            i += 1

        existing.add(nombre)

        ws = wb_out.create_sheet(nombre)

        for row in dataframe_to_rows(sub[cols_out], index=False, header=True):
            ws.append([sanitize_cell(v) for v in row])

        style_sheet(ws)

    # --- PARADAS: clave Población + calle sin número ---

    
    paradas_por_hoja = {}
    
    for hoja, subdf in [("HOSPITALES", hosp), ("FEDERACION", fed)]:
        paradas_por_hoja[hoja] = subdf["Clave_parada"].nunique()
    
    for z, subdf in resto.groupby("Z.Rep"):
        z = str(z).strip()
        if z == ".":
            z = ""
        nombre = f"ZREP_{z}"
        nombre = re.sub(r"[\\/*?:\[\]]", "_", nombre)[:31]
        paradas_por_hoja[nombre] = subdf["Clave_parada"].nunique()
    
    # Guardar paradas en un CSV auxiliar junto al Excel

    paradas_path = out_path.with_suffix(".paradas.json")
    with open(paradas_path, "w", encoding="utf-8") as f:
        json.dump(paradas_por_hoja, f, ensure_ascii=False)

    # --- GUARDAR ---
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(out_path)

# -------------------------
# MAIN
# -------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--csv")
    parser.add_argument("--reglas")
    parser.add_argument("--out")
    parser.add_argument("--delegacion", default="castellon")
    parser.add_argument("--api_key", default="")
    parser.add_argument("--coordenadas", default=None)

    args = parser.parse_args()

    csv_p = Path(args.csv)
    reglas_p = Path(args.reglas)
    out_p = Path(args.out)
    coord_p = Path(args.coordenadas) if args.coordenadas else None

    run(csv_p, reglas_p, out_p, "LLEGADAS", args.delegacion,
        api_key=args.api_key, ruta_coordenadas=coord_p)

    print(f"OK: generado {out_p}")


if __name__ == "__main__":
    main()
