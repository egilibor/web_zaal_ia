import sys
import uuid
import shutil
import tempfile
import subprocess
import datetime
import pandas as pd
from pathlib import Path

import streamlit as st
from auth import init_db, render_login, render_panel_admin, registrar_actividad
from reordenar_rutas import reordenar_excel, generar_link_pueblos, generar_links_segmentos, generar_kml
from add_resumen_unico import generar_resumen_unico
from modulo_valencia_gestores import generar_libros_gestores
from openpyxl import load_workbook

# ==========================================================
# CONFIG (debe ser el primer comando Streamlit)
# ==========================================================
st.set_page_config(page_title="Reparto determinista", layout="wide")

# ==========================================================
# BASE DE DATOS Y AUTENTICACIÓN
# ==========================================================
init_db()

if "usuario" not in st.session_state:
    render_login()
    st.stop()

usuario = st.session_state["usuario"]

# ==========================================================
# SIDEBAR — SESIÓN
# ==========================================================
with st.sidebar:
    st.markdown(f"**{usuario['nombre']}**  \n`{usuario['rol']}` · {usuario['agencia']}")
    if st.button("Cerrar sesión"):
        st.session_state.clear()
        st.rerun()
    st.markdown("---")

# ==========================================================
# WORKDIR
# ==========================================================
hora_salida = st.sidebar.time_input(
    "Hora de salida",
    value=datetime.time(8, 30)
)

if "workdir" not in st.session_state:
    st.session_state.workdir = Path(tempfile.mkdtemp(prefix="reparto_"))
    st.session_state.run_id = str(uuid.uuid4())[:8]

workdir = st.session_state.workdir

with st.sidebar:
    st.write(f"Run ID: {st.session_state.run_id}")
    st.write(f"Workdir: {workdir}")
    if st.button("Reset sesión"):
        shutil.rmtree(workdir, ignore_errors=True)
        st.session_state.workdir = Path(tempfile.mkdtemp(prefix="reparto_"))
        st.session_state.run_id = str(uuid.uuid4())[:8]
        st.rerun()
    if st.button("🗑️ Limpiar caché geocodificación"):
        from geocodificador import limpiar_cache
        limpiar_cache()
        st.success("Caché limpiada correctamente")

# ==========================================================
# CONFIG RUTAS
# ==========================================================
if "GOOGLE_MAPS_API_KEY" not in st.secrets:
    st.error("⚠️ Falta la clave GOOGLE_MAPS_API_KEY en los secrets. Contacta con el administrador.")
    st.stop()

REPO_DIR = Path(__file__).resolve().parent
SCRIPT_REPARTO = REPO_DIR / "reparto_gpt.py"
REGLAS_REPO = REPO_DIR / "Reglas_hospitales.xlsx"

# ==========================================================
# DELEGACIÓN
# - Admin:   selector en sidebar
# - Usuario: agencia asignada, sin selector
# ==========================================================
AGENCIA_MAP = {"Valencia": "valencia", "Castellon": "castellon"}

if usuario["rol"] == "admin":
    if "delegacion_activa" not in st.session_state:
        st.session_state.delegacion_activa = None

    if st.session_state.delegacion_activa is None:
        st.markdown("## Selecciona la delegación")
        st.markdown("---")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("🏙️ CASTELLÓN", use_container_width=True, type="primary"):
                st.session_state.delegacion_activa = "castellon"
                st.rerun()
        with col2:
            if st.button("🌆 VALENCIA", use_container_width=True, type="primary"):
                st.session_state.delegacion_activa = "valencia"
                st.rerun()
        st.stop()

    if st.sidebar.button("🔄 Cambiar delegación"):
        st.session_state.delegacion_activa = None
        st.session_state.pop("workdir", None)
        st.rerun()

    delegacion = st.session_state.delegacion_activa

else:
    # Usuario estándar: va directo a su agencia
    delegacion = AGENCIA_MAP.get(usuario["agencia"], "valencia")
    st.session_state.delegacion_activa = delegacion

# ==========================================================
# COORDENADAS
# ==========================================================
COORDENADAS_FILES = {
    "castellon": "Libro_de_Servicio_Castellon_con_coordenadas.xlsx",
    "valencia":  "valencia_municipios_coordenadas.xlsx",
}
COORDENADAS_REPO = REPO_DIR / COORDENADAS_FILES[delegacion]

st.title(f"Reparto determinista — {delegacion.upper()}")

# ==========================================================
# MENÚ HORIZONTAL
# Admin ve además el panel de administración
# ==========================================================
if usuario["rol"] == "admin":
    tab1, tab2, tab3, tab_refino, tab5, tab_admin = st.tabs([
        "FASE 1 · Clasificación zonas",
        "FASE 2 · Ajuste Gestores",
        "FASE 3 · Orden de Carga/Google Maps",
        "FASE 4 · Refino",
        "FASE 5 · Exportar KML",
        "⚙️ Administración",
    ])
else:
    tab1, tab2, tab3, tab_refino = st.tabs([
        "FASE 1 · Clasificación zonas",
        "FASE 2 · Ajuste Gestores",
        "FASE 3 · Orden de Carga/Google Maps",
        "FASE 4 · Refino",
    ])

# ==========================================================
# FASE 1
# ==========================================================
with tab1:

    st.subheader("Subir CSV de llegadas")

    csv_file = st.file_uploader(
        "CSV de llegadas",
        type=["csv"],
        key="fase1_csv"
    )

    if csv_file:

        input_csv = workdir / "llegadas.csv"
        input_csv.write_bytes(csv_file.getbuffer())

        (workdir / "Reglas_hospitales.xlsx").write_bytes(
            REGLAS_REPO.read_bytes()
        )

        if st.button("Generar salida.xlsx", key="fase1_btn"):

            cmd = [
                sys.executable,
                str(SCRIPT_REPARTO),
                "--csv", "llegadas.csv",
                "--reglas", "Reglas_hospitales.xlsx",
                "--out", "salida.xlsx",
                "--delegacion", delegacion,
                "--api_key", st.secrets["GOOGLE_MAPS_API_KEY"],
                "--coordenadas", str(COORDENADAS_REPO),
            ]

            with st.spinner("Ejecutando reparto_gpt.py…"):
                p = subprocess.run(
                    cmd,
                    cwd=str(workdir),
                    capture_output=True,
                    text=True,
                )

            if p.returncode != 0:
                st.error("Error en reparto_gpt.py")
                st.code(p.stderr)
            else:
                salida = workdir / "salida.xlsx"
                if salida.exists():
                    # Crear hoja ALMACEN antes de generar resumen para que aparezca en él
                    from openpyxl import load_workbook as _lw_f1
                    _wb_f1 = _lw_f1(salida)
                    if "ALMACEN" not in _wb_f1.sheetnames:
                        _ref_hoja = next(
                            (h for h in ["HOSPITALES", "FEDERACION"] + sorted([s for s in _wb_f1.sheetnames if s.startswith("ZREP_")])
                             if h in _wb_f1.sheetnames),
                            None
                        )
                        _idx_hosp = _wb_f1.sheetnames.index("HOSPITALES") if "HOSPITALES" in _wb_f1.sheetnames else 0
                        _ws_alm = _wb_f1.create_sheet(title="ALMACEN", index=_idx_hosp)
                        if _ref_hoja:
                            _cabeceras = [c.value for c in next(_wb_f1[_ref_hoja].iter_rows(min_row=1, max_row=1))]
                            _ws_alm.append(_cabeceras)
                        _wb_f1.save(salida)
                    generar_resumen_unico(str(salida))
                    registrar_actividad(usuario["id"], usuario["nombre"], delegacion, "Fase 1 - Clasificación zonas")
                    st.success("Archivo generado correctamente")

                    st.download_button(
                        "Descargar salida.xlsx",
                        data=salida.read_bytes(),
                        file_name="salida.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                    # Excel por gestor solo para Valencia
                    if delegacion == "valencia":
                        ruta_asignacion = REPO_DIR / "gestor_zonas.xlsx"
                        resultado_gestores = generar_libros_gestores(
                            ruta_excel_final=str(salida),
                            ruta_asignacion=str(ruta_asignacion),
                            carpeta_salida=str(workdir)
                        )
                        if resultado_gestores["ok"]:
                            st.markdown("---")
                            st.subheader("Excel por gestor de tráfico")
                            for gestor, ruta_archivo in resultado_gestores["archivos_generados"].items():
                                ruta = Path(ruta_archivo)
                                st.download_button(
                                    label=f"Descargar Excel {gestor}",
                                    data=ruta.read_bytes(),
                                    file_name=ruta.name,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                )
                        else:
                            for e in resultado_gestores["errores"]:
                                st.warning(e)
    else:
        st.info("Sube un CSV para habilitar la ejecución.")

# ==========================================================
# FASE 2
# ==========================================================
with tab2:

    st.subheader("Ajuste manual de expediciones")

    excel_ajuste = st.file_uploader(
        "Subir salida.xlsx (generado en Fase 1)",
        type=["xlsx"],
        key="fase2_ajuste_excel"
    )

    if excel_ajuste:

        file_id = excel_ajuste.name + str(excel_ajuste.size)
        if st.session_state.get("ajuste_file_id") != file_id:
            ajuste_path = workdir / "ajuste_entrada.xlsx"
            ajuste_path.write_bytes(excel_ajuste.getbuffer())
            wb = load_workbook(ajuste_path)
            st.session_state["ajuste_wb"] = wb
            st.session_state["ajuste_file_id"] = file_id

        wb = st.session_state["ajuste_wb"]
        hojas_disponibles = wb.sheetnames
        hojas_operativas = [
            h for h in hojas_disponibles
            if h.startswith("ZREP_") or h in ("ALMACEN", "HOSPITALES", "FEDERACION")
        ]

        if not hojas_operativas:
            st.warning("No se encontraron hojas operativas (ZREP_, HOSPITALES, FEDERACION).")
        else:
            col_iz, col_der = st.columns(2)

            with col_iz:
                hoja_origen = st.selectbox("Hoja origen", hojas_operativas, key="hoja_origen")
            with col_der:
                hojas_destino = [h for h in hojas_operativas if h != hoja_origen]
                hoja_destino = st.selectbox("Hoja destino", hojas_destino, key="hoja_destino")

            def ws_to_df(wb, nombre):
                ws = wb[nombre]
                datos = [r for r in ws.values if not (r[0] == "← RESUMEN")]
                if not datos:
                    return pd.DataFrame()
                return pd.DataFrame(datos[1:], columns=datos[0])

            df_origen = ws_to_df(wb, hoja_origen)
            df_destino = ws_to_df(wb, hoja_destino)

            columnas_mostrar = [
                c for c in ["Exp", "Ref.", "Consignatario", "Población", "Kgs", "B.Doc", "Obs.", "AmpFtiI", "ObsClt", "F.Teo.Entr.", "Dirección"]
                if c in df_origen.columns
            ]

            with st.expander(f"Ver hoja destino: {hoja_destino} ({len(df_destino)} expediciones)"):
                st.dataframe(df_destino[columnas_mostrar] if columnas_mostrar else df_destino, use_container_width=True)

            total_btos = int(df_origen["Bultos"].apply(pd.to_numeric, errors="coerce").sum()) if "Bultos" in df_origen.columns else 0
            total_kgs = df_origen["Kgs"].apply(pd.to_numeric, errors="coerce").sum() if "Kgs" in df_origen.columns else 0
            st.markdown(f"**{hoja_origen}** — {len(df_origen)} expediciones · {total_btos} btos · {total_kgs:.0f} kg")

            sel_key = f"sel_df_{hoja_origen}"
            if sel_key not in st.session_state or len(st.session_state[sel_key]) != len(df_origen):
                st.session_state[sel_key] = [False] * len(df_origen)

            master_anterior = st.session_state.get("chk_master_prev", False)

            with st.form(key=f"form_{hoja_origen}"):
                master_actual = st.checkbox("Seleccionar todas", key="chk_master")

                df_editor = df_origen[columnas_mostrar].copy()
                df_editor.insert(0, "✓", pd.Series(st.session_state[sel_key], dtype=bool))

                edited = st.data_editor(
                    df_editor,
                    use_container_width=True,
                    hide_index=True,
                    column_config={"✓": st.column_config.CheckboxColumn("✓", default=False)},
                    disabled=columnas_mostrar,
                    key=f"editor_{hoja_origen}",
                )

                indices_seleccionados = [df_origen.index[i] for i, sel in enumerate(edited["✓"]) if sel]
                st.markdown(f"*{len(indices_seleccionados)} expedición(es) seleccionada(s)*")

                accion = st.radio(
                    "¿Qué hacer con las expediciones seleccionadas?",
                    ["Mover a otra ruta", "Mover a 2º reparto", "Mover a ALMACEN"],
                    key="radio_accion",
                    horizontal=True
                )

                submitted = st.form_submit_button("Ejecutar acción")

                if submitted:
                    indices_seleccionados = [df_origen.index[i] for i, sel in enumerate(edited["✓"]) if sel]
                    if not indices_seleccionados:
                        st.warning("Selecciona al menos una expedición antes de ejecutar una acción.")
                    else:
                        st.session_state[sel_key] = list(edited["✓"])
                        if master_actual and not master_anterior:
                            st.session_state[sel_key] = [True] * len(df_origen)
                        elif not master_actual and master_anterior:
                            st.session_state[sel_key] = [False] * len(df_origen)
                        st.session_state["chk_master_prev"] = master_actual

                        if accion == "Mover a otra ruta":
                            from openpyxl.utils.dataframe import dataframe_to_rows
                            df_src = ws_to_df(wb, hoja_origen)
                            df_dst = ws_to_df(wb, hoja_destino)
                            filas_a_mover = df_src.loc[indices_seleccionados]
                            df_src_nuevo = df_src.drop(index=indices_seleccionados).reset_index(drop=True)
                            df_dst_nuevo = pd.concat([df_dst, filas_a_mover], ignore_index=True)
                            for nombre_hoja, df_nuevo in [(hoja_origen, df_src_nuevo), (hoja_destino, df_dst_nuevo)]:
                                ws = wb[nombre_hoja]
                                ws.delete_rows(1, ws.max_row)
                                for r in dataframe_to_rows(df_nuevo, index=False, header=True):
                                    ws.append(r)
                            st.session_state["ajuste_wb"] = wb
                            registrar_actividad(usuario["id"], usuario["nombre"], delegacion, "Fase 2 - Ajuste Gestores")
                            st.success(f"{len(indices_seleccionados)} expedición(es) movidas de '{hoja_origen}' a '{hoja_destino}'")
                            st.rerun()

                        elif accion == "Mover a ALMACEN":
                            from openpyxl.utils.dataframe import dataframe_to_rows
                            df_src = ws_to_df(wb, hoja_origen)
                            filas_alm = df_src.loc[indices_seleccionados]
                            df_src_nuevo = df_src.drop(index=indices_seleccionados).reset_index(drop=True)
                            ws_a = wb[hoja_origen]
                            ws_a.delete_rows(1, ws_a.max_row)
                            for r in dataframe_to_rows(df_src_nuevo, index=False, header=True):
                                ws_a.append(r)
                            if "ALMACEN" not in wb.sheetnames:
                                ws_alm = wb.create_sheet(title="ALMACEN")
                                for r in dataframe_to_rows(filas_alm, index=False, header=True):
                                    ws_alm.append(r)
                            else:
                                df_alm_actual = ws_to_df(wb, "ALMACEN")
                                df_alm_nuevo = pd.concat([df_alm_actual, filas_alm], ignore_index=True)
                                ws_alm = wb["ALMACEN"]
                                ws_alm.delete_rows(1, ws_alm.max_row)
                                for r in dataframe_to_rows(df_alm_nuevo, index=False, header=True):
                                    ws_alm.append(r)
                            st.session_state["ajuste_wb"] = wb
                            registrar_actividad(usuario["id"], usuario["nombre"], delegacion, "Fase 2 - Mover a ALMACEN")
                            st.success(f"{len(indices_seleccionados)} expedición(es) movidas a ALMACEN")
                            st.rerun()

                        elif accion == "Mover a 2º reparto":
                            from openpyxl.utils.dataframe import dataframe_to_rows
                            nombre_b = hoja_origen + "_B"
                            df_src = ws_to_df(wb, hoja_origen)
                            filas_b = df_src.loc[indices_seleccionados]
                            filas_a = df_src.drop(index=indices_seleccionados).reset_index(drop=True)
                            ws_a = wb[hoja_origen]
                            ws_a.delete_rows(1, ws_a.max_row)
                            for r in dataframe_to_rows(filas_a, index=False, header=True):
                                ws_a.append(r)
                            if nombre_b in wb.sheetnames:
                                ws_b = wb[nombre_b]
                                ws_b.delete_rows(1, ws_b.max_row)
                            else:
                                ws_b = wb.create_sheet(title=nombre_b)
                            for r in dataframe_to_rows(filas_b, index=False, header=True):
                                ws_b.append(r)
                            from openpyxl import Workbook as WB2
                            wb_nuevo = WB2()
                            wb_nuevo.remove(wb_nuevo.active)
                            for nombre_hoja in wb.sheetnames:
                                ws_orig = wb[nombre_hoja]
                                ws_nuevo = wb_nuevo.create_sheet(title=nombre_hoja)
                                for row in ws_orig.iter_rows(values_only=True):
                                    ws_nuevo.append(list(row) if row else [])
                            st.session_state["ajuste_wb"] = wb_nuevo
                            ajuste_salida = workdir / "ajuste_salida.xlsx"
                            wb_nuevo.save(ajuste_salida)
                            from add_resumen_unico import generar_resumen_unico
                            generar_resumen_unico(str(ajuste_salida))
                            wb_actualizado = load_workbook(ajuste_salida)
                            st.session_state["ajuste_wb"] = wb_actualizado
                            registrar_actividad(usuario["id"], usuario["nombre"], delegacion, "Fase 2 - 2º reparto")
                            st.success(f"2º reparto creado: '{nombre_b}' con {len(filas_b)} expedición(es)")
                            st.rerun()

            indices_seleccionados = [df_origen.index[i] for i, sel in enumerate(st.session_state[sel_key]) if sel]
            st.markdown(f"*{len(indices_seleccionados)} expedición(es) seleccionada(s)*")

            if indices_seleccionados:
                accion = st.radio(
                    "¿Qué hacer con las expediciones seleccionadas?",
                    ["Mover a otra ruta", "Mover a 2º reparto", "Mover a ALMACEN"],
                    key="radio_accion",
                    horizontal=True
                )

                if accion == "Mover a otra ruta":
                    if st.button(f"Mover {len(indices_seleccionados)} exp. → {hoja_destino}", key="btn_mover"):
                        from openpyxl.utils.dataframe import dataframe_to_rows
                        df_src = ws_to_df(wb, hoja_origen)
                        df_dst = ws_to_df(wb, hoja_destino)
                        filas_a_mover = df_src.loc[indices_seleccionados]
                        df_src_nuevo = df_src.drop(index=indices_seleccionados).reset_index(drop=True)
                        df_dst_nuevo = pd.concat([df_dst, filas_a_mover], ignore_index=True)
                        for nombre_hoja, df_nuevo in [(hoja_origen, df_src_nuevo), (hoja_destino, df_dst_nuevo)]:
                            ws = wb[nombre_hoja]
                            ws.delete_rows(1, ws.max_row)
                            for r in dataframe_to_rows(df_nuevo, index=False, header=True):
                                ws.append(r)
                        st.session_state["ajuste_wb"] = wb
                        registrar_actividad(usuario["id"], usuario["nombre"], delegacion, "Fase 2 - Ajuste Gestores")
                        st.success(f"{len(indices_seleccionados)} expedición(es) movidas de '{hoja_origen}' a '{hoja_destino}'")
                        st.rerun()

                elif accion == "Mover a ALMACEN":
                    if st.button(f"Mover {len(indices_seleccionados)} exp. → ALMACEN", key="btn_almacen"):
                        from openpyxl.utils.dataframe import dataframe_to_rows
                        df_src = ws_to_df(wb, hoja_origen)
                        filas_alm = df_src.loc[indices_seleccionados]
                        df_src_nuevo = df_src.drop(index=indices_seleccionados).reset_index(drop=True)
                        ws_a = wb[hoja_origen]
                        ws_a.delete_rows(1, ws_a.max_row)
                        for r in dataframe_to_rows(df_src_nuevo, index=False, header=True):
                            ws_a.append(r)
                        if "ALMACEN" not in wb.sheetnames:
                            ws_alm = wb.create_sheet(title="ALMACEN")
                            for r in dataframe_to_rows(filas_alm, index=False, header=True):
                                ws_alm.append(r)
                        else:
                            df_alm_actual = ws_to_df(wb, "ALMACEN")
                            df_alm_nuevo = pd.concat([df_alm_actual, filas_alm], ignore_index=True)
                            ws_alm = wb["ALMACEN"]
                            ws_alm.delete_rows(1, ws_alm.max_row)
                            for r in dataframe_to_rows(df_alm_nuevo, index=False, header=True):
                                ws_alm.append(r)
                        st.session_state["ajuste_wb"] = wb
                        registrar_actividad(usuario["id"], usuario["nombre"], delegacion, "Fase 2 - Mover a ALMACEN")
                        st.success(f"{len(indices_seleccionados)} expedición(es) movidas a ALMACEN")
                        st.rerun()

                elif accion == "Mover a 2º reparto":
                    if st.button(f"Mover {len(indices_seleccionados)} exp. → 2º reparto", key="btn_segundo_reparto"):
                        from openpyxl.utils.dataframe import dataframe_to_rows
                        nombre_b = hoja_origen + "_B"
                        df_src = ws_to_df(wb, hoja_origen)
                        filas_b = df_src.loc[indices_seleccionados]
                        filas_a = df_src.drop(index=indices_seleccionados).reset_index(drop=True)
                        ws_a = wb[hoja_origen]
                        ws_a.delete_rows(1, ws_a.max_row)
                        for r in dataframe_to_rows(filas_a, index=False, header=True):
                            ws_a.append(r)
                        if nombre_b in wb.sheetnames:
                            ws_b = wb[nombre_b]
                            ws_b.delete_rows(1, ws_b.max_row)
                        else:
                            ws_b = wb.create_sheet(title=nombre_b)
                        for r in dataframe_to_rows(filas_b, index=False, header=True):
                            ws_b.append(r)
                        from openpyxl import Workbook as WB2
                        wb_nuevo = WB2()
                        wb_nuevo.remove(wb_nuevo.active)
                        for nombre_hoja in wb.sheetnames:
                            ws_orig = wb[nombre_hoja]
                            ws_nuevo = wb_nuevo.create_sheet(title=nombre_hoja)
                            for row in ws_orig.iter_rows(values_only=True):
                                ws_nuevo.append(list(row) if row else [])
                        st.session_state["ajuste_wb"] = wb_nuevo
                        ajuste_salida = workdir / "ajuste_salida.xlsx"
                        wb_nuevo.save(ajuste_salida)
                        from add_resumen_unico import generar_resumen_unico
                        generar_resumen_unico(str(ajuste_salida))
                        wb_actualizado = load_workbook(ajuste_salida)
                        st.session_state["ajuste_wb"] = wb_actualizado
                        registrar_actividad(usuario["id"], usuario["nombre"], delegacion, "Fase 2 - 2º reparto")
                        st.success(f"2º reparto creado: '{nombre_b}' con {len(filas_b)} expedición(es)")
                        st.rerun()

            ajuste_salida = workdir / "ajuste_salida.xlsx"
            wb.save(ajuste_salida)
            st.download_button(
                "⬇️ Descargar Excel modificado",
                data=ajuste_salida.read_bytes(),
                file_name="ajuste_salida.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="btn_descarga_ajuste"
            )

    else:
        st.info("Sube el salida.xlsx de la Fase 1 para hacer ajustes manuales.")

# ==========================================================
# FASE 3
# ==========================================================
with tab3:

    st.subheader("Reordenar rutas existentes")

    archivo_excel = st.file_uploader(
        "Subir salida.xlsx modificado",
        type=["xlsx"],
        key="fase2_excel"
    )

    if archivo_excel:

        input_path  = workdir / "entrada_fase2.xlsx"
        output_path = workdir / "salida_reordenada.xlsx"

        input_path.write_bytes(archivo_excel.getbuffer())

        if st.button("Reordenar rutas", key="fase2_btn"):

            try:

                if delegacion == "valencia":
                    lat_origen = 39.44069
                    lon_origen = -0.42589
                else:
                    lat_origen = 39.804106
                    lon_origen = -0.217351

                paradas = reordenar_excel(
                    input_path,
                    output_path,
                    COORDENADAS_REPO,
                    lat_origen,
                    lon_origen,
                    api_key=st.secrets["GOOGLE_MAPS_API_KEY"],
                    delegacion=delegacion,
                    hora_salida=hora_salida,
                )

                generar_resumen_unico(str(output_path), paradas_por_hoja=paradas)

                if output_path.exists():
                    registrar_actividad(usuario["id"], usuario["nombre"], delegacion, "Fase 3 - Orden de Carga")
                    st.success("Rutas reordenadas correctamente")

                    st.download_button(
                        "Descargar salida_reordenada.xlsx",
                        data=output_path.read_bytes(),
                        file_name="salida_reordenada.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                else:
                    st.error("No se generó el archivo reordenado.")

            except Exception as e:
                import traceback
                st.error(f"Error en reordenación: {e}")
                st.code(traceback.format_exc())

    else:
        st.info("Sube el archivo para activar la reordenación.")

# ==========================================================
# FASE DE REFINO
# ==========================================================
with tab_refino:

    st.subheader("Ajuste fino del orden de expediciones")

    archivo_refino = st.file_uploader(
        "Subir salida_reordenada.xlsx",
        type=["xlsx"],
        key="refino_excel"
    )

    if archivo_refino:

        _file_id_refino = archivo_refino.name + str(archivo_refino.size)
        if st.session_state.get("refino_file_id") != _file_id_refino:
            _refino_path = workdir / "refino_entrada.xlsx"
            _refino_path.write_bytes(archivo_refino.getbuffer())
            st.session_state["refino_file_id"] = _file_id_refino
            st.session_state["refino_path"] = str(_refino_path)
            for _k in [_k for _k in st.session_state if _k.startswith("refino_orden_")]:
                del st.session_state[_k]

        _refino_path = Path(st.session_state["refino_path"])

        _wb_tmp = load_workbook(_refino_path, read_only=True)
        _hojas_refino = [
            h for h in _wb_tmp.sheetnames
            if h.startswith("ZREP_") or h in ("ALMACEN", "HOSPITALES", "FEDERACION")
        ]
        _wb_tmp.close()

        if not _hojas_refino:
            st.warning("No se encontraron hojas operativas (ZREP_, HOSPITALES, FEDERACION).")
        else:
            _hoja_refino = st.selectbox("Seleccionar hoja", _hojas_refino, key="refino_hoja")

            _wb_r = load_workbook(_refino_path, read_only=True, data_only=True)
            _ws_r = _wb_r[_hoja_refino]
            _all_rows = [tuple(c.value for c in row) for row in _ws_r.iter_rows()]
            _wb_r.close()

            _hdr_idx = None
            for _i, _row in enumerate(_all_rows):
                if _row and "Exp" in _row:
                    _hdr_idx = _i
                    break

            if _hdr_idx is None:
                st.error("No se encontró la cabecera (columna Exp) en la hoja seleccionada.")
            else:
                _headers = list(_all_rows[_hdr_idx])
                _data_rows = _all_rows[_hdr_idx + 1:]

                _orden_key = f"refino_orden_{_hoja_refino}"
                if _orden_key not in st.session_state or len(st.session_state[_orden_key]) != len(_data_rows):
                    st.session_state[_orden_key] = list(range(len(_data_rows)))

                    st.session_state[_orden_key] = list(range(len(_data_rows)))
                    st.session_state.pop(f"bloque_ini_{_hoja_refino}", None)
                    st.session_state.pop(f"bloque_fin_{_hoja_refino}", None)
                    st.session_state.pop(f"bloque_dest_{_hoja_refino}", None)

                _orden = st.session_state[_orden_key]
                _col_idx = {h: i for i, h in enumerate(_headers) if h is not None}
                _cols_mostrar = [c for c in ["Exp", "Ref.", "Población", "Dirección", "Consignatario"] if c in _col_idx]

                st.markdown(f"**{_hoja_refino}** — {len(_data_rows)} expediciones")
                st.markdown("---")

                from streamlit_sortables import sort_items as _sort_items

                st.markdown("""
                <style>
                [data-testid="stCustomComponentV1"] iframe {
                    font-family: 'Courier New', Courier, monospace !important;
                }
                .sortable-item, .draggable-item, [class*="sortable"] {
                    font-family: 'Courier New', Courier, monospace !important;
                }
                </style>
                """, unsafe_allow_html=True)

                if f"sortable_version_{_hoja_refino}" not in st.session_state:
                    st.session_state[f"sortable_version_{_hoja_refino}"] = 0

                _almacen_key = f"almacen_items_{_hoja_refino}"
                if _almacen_key not in st.session_state:
                    st.session_state[_almacen_key] = []

                _ANCHOS_F4_R1 = {
                    "Exp": 15, "Consignatario": 30, "Población": 30,
                    "Kgs": 7, "B.Doc": 5, "Compromiso": 5, "F.Max.Ent": 12, "F.Teo.Entr.": 10, "Dirección": 30,
                }
                _ANCHOS_F4 = _ANCHOS_F4_R1
                _OBS_COLS = ["Obs.", "AmpFtiI", "ObsClt"]
                _INDENT = "     "

                def _trunc(val, ancho):
                    s = "" if val is None or str(val).strip() in ("", "nan") else str(val).strip()
                    if len(s) > ancho:
                        s = s[:ancho - 1] + "…"
                    return s + "-" * (ancho - len(s))

                _header_parts = []
                for _col_name, _ancho in _ANCHOS_F4_R1.items():
                    if _col_name in _col_idx:
                        _h = _col_name if len(_col_name) <= _ancho else _col_name[:_ancho - 1] + "…"
                        _header_parts.append(_h + "-" * (_ancho - len(_h)))
                _hdr_r1 = "  · ".join(_header_parts)
                st.markdown(
                    f'<div style="font-family:\'Courier New\',Courier,monospace;font-size:0.85rem;'
                    f'background-color:#444;color:#ff0;padding:4px 14px;border-radius:3px;'
                    f'margin-bottom:4px;white-space:pre;letter-spacing:0;">'
                    f'##  {_hdr_r1}</div>',
                    unsafe_allow_html=True,
                )

                _items_labels = []
                for _pos, _orig_idx in enumerate(_orden):
                    _row_data = _data_rows[_orig_idx]
                    _partes_r1 = []
                    for _col_name, _ancho in _ANCHOS_F4_R1.items():
                        if _col_name in _col_idx:
                            _v = _row_data[_col_idx[_col_name]]
                            _partes_r1.append(_trunc(_v, _ancho))
                    _obs_parts = []
                    for _col_name in _OBS_COLS:
                        if _col_name in _col_idx:
                            _v = _row_data[_col_idx[_col_name]]
                            _s = "" if _v is None or str(_v).strip() in ("", "nan") else str(_v).strip()
                            if _s:
                                _obs_parts.append(_s)
                    _obs_str = ", ".join(_obs_parts) if _obs_parts else ""
                    _label = f"{_pos + 1}. " + " · ".join(_partes_r1)
                    if _obs_str:
                        _label += f"\n{_INDENT}{_obs_str}"
                    _label += f"  [{_orig_idx}]"
                    _items_labels.append(_label)

                _version = st.session_state.get(f"sortable_version_{_hoja_refino}", 0)
                _containers_in = [
                    {"header": _hoja_refino, "items": _items_labels},
                    {"header": "ALMACEN", "items": st.session_state[_almacen_key]},
                ]
                _containers_out = _sort_items(
                    _containers_in,
                    multi_containers=True,
                    direction="vertical",
                    key=f"sortable_{_hoja_refino}_{_version}",
                    custom_style=".sortable-item { background-color: #f2f2f2 !important; color: #000000 !important; border: 1px solid #aaaaaa !important; border-radius: 4px !important; margin-bottom: 4px !important; font-weight: bold !important; }",
                )

                _sorted_labels = _containers_out[0]["items"]
                _almacen_labels = _containers_out[1]["items"]

                _nuevo_orden = [int(_lbl.split("[")[-1].rstrip("]")) for _lbl in _sorted_labels]
                if _nuevo_orden != st.session_state.get(_orden_key) or _almacen_labels != st.session_state[_almacen_key]:
                    st.session_state[_orden_key] = _nuevo_orden
                    st.session_state[_almacen_key] = _almacen_labels

                st.markdown("---")
                st.markdown("**Mover bloque de expediciones**")
                _n_exp = len(st.session_state[_orden_key])
                if _n_exp == 0:
                    st.info("Esta hoja no tiene expediciones.")
                    st.stop()
                _bc1, _bc2, _bc3, _bc4 = st.columns([1, 1, 1, 1])
                with _bc1:
                    _bloque_ini = st.number_input(
                        "Posición inicial", min_value=1, max_value=_n_exp,
                        value=1, step=1, key=f"bloque_ini_{_hoja_refino}"
                    )
                with _bc2:
                    _bloque_fin = st.number_input(
                        "Posición final", min_value=1, max_value=_n_exp,
                        value=1, step=1, key=f"bloque_fin_{_hoja_refino}"
                    )
                with _bc3:
                    _bloque_dest = st.number_input(
                        "Insertar en posición", min_value=1, max_value=_n_exp,
                        value=1, step=1, key=f"bloque_dest_{_hoja_refino}"
                    )
                with _bc4:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button("Mover bloque", key=f"mover_bloque_{_hoja_refino}"):
                        _ini = int(_bloque_ini)
                        _fin = int(_bloque_fin)
                        _dest = int(_bloque_dest)
                        if _ini > _fin:
                            st.error("La posición inicial debe ser menor o igual a la final.")
                        elif _ini == _dest:
                            st.info("El bloque ya está en esa posición.")
                        else:
                            _ord_actual = list(st.session_state[_orden_key])
                            _bloque = _ord_actual[_ini - 1:_fin]
                            _resto = _ord_actual[:_ini - 1] + _ord_actual[_fin:]
                            _insert_idx = min(_dest - 1, len(_resto))
                            _ord_nuevo = _resto[:_insert_idx] + _bloque + _resto[_insert_idx:]
                            st.session_state[_orden_key] = _ord_nuevo
                            st.session_state[f"sortable_version_{_hoja_refino}"] += 1
                            st.rerun()

                st.markdown("---")

                if st.button("Guardar y regenerar rutas", key="refino_guardar", type="primary"):
                    try:
                        from openpyxl.styles import Font, PatternFill

                        _wb_save = load_workbook(_refino_path)
                        _ws_save = _wb_save[_hoja_refino]

                        _save_rows = list(_ws_save.iter_rows(values_only=True))
                        _save_hdr = None
                        for _si, _sr in enumerate(_save_rows):
                            if _sr and "Exp" in _sr:
                                _save_hdr = _si + 1  # 1-indexed
                                break

                        if _save_hdr is None:
                            st.error("Error: no se encontró la cabecera al guardar.")
                        else:
                            _save_header_vals = list(_save_rows[_save_hdr - 1])
                            _save_data_vals = _save_rows[_save_hdr:]
                            _data_nuevo = [_save_data_vals[i] for i in _orden]
                            _n_nav_save = _save_hdr - 1

                            # Clear images (anchors become invalid after row reorder)
                            _ws_save._images = []

                            # Delete old data rows and write reordered values
                            if _ws_save.max_row > _save_hdr:
                                _ws_save.delete_rows(_save_hdr + 1, _ws_save.max_row - _save_hdr)

                            _azul_claro = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
                            _parada_col = _save_header_vals.index("Parada") if "Parada" in _save_header_vals else None

                            for _ri, _row_vals in enumerate(_data_nuevo):
                                _excel_row = _save_hdr + 1 + _ri
                                for _ci_v, _v in enumerate(list(_row_vals) if _row_vals else []):
                                    _ws_save.cell(row=_excel_row, column=_ci_v + 1).value = _v
                                if _parada_col is not None and _row_vals:
                                    try:
                                        if int(_row_vals[_parada_col]) % 2 != 0:
                                            for _ci_v in range(len(_row_vals)):
                                                _ws_save.cell(row=_excel_row, column=_ci_v + 1).fill = _azul_claro
                                    except (TypeError, ValueError):
                                        pass

                            # Regenerate Google Maps links
                            _col_lat = _save_header_vals.index("Latitud") if "Latitud" in _save_header_vals else None
                            _col_lon = _save_header_vals.index("Longitud") if "Longitud" in _save_header_vals else None

                            if _col_lat is not None and _col_lon is not None:
                                _df_refino = pd.DataFrame(_data_nuevo, columns=_save_header_vals)

                                if delegacion == "valencia":
                                    _lat_orig, _lon_orig = 39.44069, -0.42589
                                else:
                                    _lat_orig, _lon_orig = 39.804106, -0.217351

                                _link_completo = generar_link_pueblos(_df_refino, _lat_orig, _lon_orig)
                                _segmentos = generar_links_segmentos(_df_refino, _lat_orig, _lon_orig)

                                # Update navigation rows
                                _ws_save.cell(row=1, column=2).value = _link_completo
                                _ws_save.cell(row=1, column=2).font = Font(color="0000FF", underline="single")
                                for _si2, _slink in enumerate(_segmentos):
                                    _seg_row = 2 + _si2
                                    if _seg_row <= _n_nav_save:
                                        _ws_save.cell(row=_seg_row, column=2).value = _slink
                                        _ws_save.cell(row=_seg_row, column=2).font = Font(color="0000FF", underline="single")

                            # Add/update REFINO history sheet
                            _now = datetime.datetime.now()
                            if "REFINO" not in _wb_save.sheetnames:
                                _ws_refino_hist = _wb_save.create_sheet("REFINO")
                                _ws_refino_hist.append(["Fecha", "Hora", "Usuario", "Hoja modificada"])
                            else:
                                _ws_refino_hist = _wb_save["REFINO"]

                            _ws_refino_hist.append([
                                _now.strftime("%Y-%m-%d"),
                                _now.strftime("%H:%M:%S"),
                                usuario["nombre"],
                                _hoja_refino,
                            ])

                            # Mover expediciones del contenedor ALMACEN a la hoja ALMACEN
                            _almacen_labels_save = st.session_state.get(_almacen_key, [])
                            if _almacen_labels_save:
                                _almacen_idxs = [int(_lbl.split("[")[-1].rstrip("]")) for _lbl in _almacen_labels_save]
                                # Usar _save_data_vals (mismo workbook que _save_header_vals) para garantizar alineación
                                _filas_almacen = [list(_save_data_vals[_ai]) for _ai in _almacen_idxs]
                                if "ALMACEN" not in _wb_save.sheetnames:
                                    _ws_alm_save = _wb_save.create_sheet(title="ALMACEN")
                                    _ws_alm_save.append(list(_save_header_vals))
                                else:
                                    _ws_alm_save = _wb_save["ALMACEN"]
                                    _alm_rows_all = list(_ws_alm_save.iter_rows(values_only=True))
                                    _alm_hdr_row = next((i + 1 for i, r in enumerate(_alm_rows_all) if r and "Exp" in r), None)
                                    if _alm_hdr_row:
                                        for _ci_h, _hv in enumerate(_save_header_vals, 1):
                                            _ws_alm_save.cell(row=_alm_hdr_row, column=_ci_h).value = _hv
                                for _fila_alm in _filas_almacen:
                                    _ws_alm_save.append(_fila_alm)
                                st.session_state[_almacen_key] = []

                            _wb_save.save(_refino_path)

                            registrar_actividad(
                                usuario["id"], usuario["nombre"], delegacion,
                                f"Fase de Refino - {_hoja_refino}"
                            )

                            st.success("Rutas regeneradas correctamente.")
                            st.download_button(
                                "Descargar salida_reordenada.xlsx actualizado",
                                data=_refino_path.read_bytes(),
                                file_name="salida_reordenada.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="refino_download"
                            )

                    except Exception as _e:
                        import traceback
                        st.error(f"Error al guardar: {_e}")
                        st.code(traceback.format_exc())

    else:
        st.info("Sube el salida_reordenada.xlsx de la Fase 3 para ajustar el orden.")

# ==========================================================
# FASE 5 · EXPORTAR KML (solo admin)
# ==========================================================
if usuario["rol"] == "admin":
    with tab5:

        st.subheader("Exportar ruta en formato KML")

        archivo_kml = st.file_uploader(
            "Subir salida_reordenada.xlsx",
            type=["xlsx"],
            key="fase5_excel"
        )

        if archivo_kml:

            _kml_file_id = archivo_kml.name + str(archivo_kml.size)
            if st.session_state.get("kml_file_id") != _kml_file_id:
                _kml_path = workdir / "kml_entrada.xlsx"
                _kml_path.write_bytes(archivo_kml.getbuffer())
                st.session_state["kml_file_id"] = _kml_file_id
                st.session_state["kml_path"] = str(_kml_path)

            _kml_path = Path(st.session_state["kml_path"])

            _wb_kml = load_workbook(_kml_path, read_only=True)
            _hojas_kml = [
                h for h in _wb_kml.sheetnames
                if h.startswith("ZREP_") or h in ("HOSPITALES", "FEDERACION")
            ]
            _wb_kml.close()

            if not _hojas_kml:
                st.warning("No se encontraron hojas operativas (ZREP_, HOSPITALES, FEDERACION).")
            else:
                _hoja_kml = st.selectbox("Selecciona la zona", _hojas_kml, key="kml_hoja_sel")

                if delegacion == "valencia":
                    _lat_kml, _lon_kml = 39.44069, -0.42589
                else:
                    _lat_kml, _lon_kml = 39.804106, -0.217351

                _df_kml = pd.read_excel(_kml_path, sheet_name=_hoja_kml, header=None)

                _hdr_row = None
                for _i, _row in _df_kml.iterrows():
                    if "Exp" in _row.values:
                        _hdr_row = _i
                        break

                if _hdr_row is None:
                    st.error("No se encontró la cabecera en la hoja seleccionada.")
                else:
                    _df_kml.columns = _df_kml.iloc[_hdr_row]
                    _df_kml = _df_kml.iloc[_hdr_row + 1:].reset_index(drop=True)

                    _kml_str = generar_kml(_df_kml, _hoja_kml, _lat_kml, _lon_kml)

                    st.download_button(
                        f"⬇️ Descargar {_hoja_kml}.kml",
                        data=_kml_str.encode("utf-8"),
                        file_name=f"{_hoja_kml}.kml",
                        mime="application/vnd.google-earth.kml+xml",
                        key="kml_download_btn"
                    )

        else:
            st.info("Sube el salida_reordenada.xlsx para generar el KML.")

# ==========================================================
# PANEL DE ADMINISTRACIÓN (solo admin)
# ==========================================================
if usuario["rol"] == "admin":
    with tab_admin:
        render_panel_admin()